import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl

wb = openpyxl.load_workbook('C:/Users/Excalıbur/Downloads/vizemo_lead_tracking.xlsx', data_only=True)

def clean(v):
    if v is None: return ''
    s = str(v).strip()
    if s.startswith('p:'): s = s[2:].strip()
    if s in ['None', 'none']: return ''
    if ' 00:00:00' in s: s = s.replace(' 00:00:00', '')
    s = s.replace('"', "'")
    s = s.replace('\n', ' ').replace('\r', '')
    bs = chr(92)
    s = s.replace(bs, '/')
    return s

def map_durum(d):
    d = str(d).lower()
    if 'olumsuz' in d: return 'Olumsuz'
    if 'tamamland' in d or 'onaylan' in d: return 'Tamamlandı'
    if any(x in d for x in ['beklem','toplan','evrak','pasaport','randevu']): return 'Beklemede'
    return 'Yeni Lead'

def map_vize(ulke, vize):
    combined = (str(vize) + ' ' + str(ulke)).lower()
    if 'oturum' in combined: return 'İspanya Oturum'
    if 'ingiltere' in combined: return 'İngiltere'
    if 'amerika' in combined or 'usa' in combined: return 'Amerika'
    if 'schengen' in combined: return 'Schengen'
    v = str(vize).strip()
    u = str(ulke).strip()
    return v if v else (u if u else 'Diğer')

idx = 1000

def read_sheet(ws, sorumlu, sehir, nc, pc, ec, kc, uc, vc, sc, dc, evc, ntc, sr=6):
    global idx
    rows = []
    for row in ws.iter_rows(min_row=sr, max_row=ws.max_row, values_only=True):
        if not row[nc]: continue
        ad = clean(row[nc])
        if not ad or 'Ad Soyad' in ad: continue
        idx += 1
        def g(col):
            if col is None: return ''
            if len(row) <= col: return ''
            return clean(row[col])
        ulke_v = g(uc)
        vize_v = g(vc)
        rows.append(
            f'  {{ id: "{idx}", ad: "{ad}", telefon: "{g(pc)}", email: "{g(ec)}",'
            f' vize: "{map_vize(ulke_v, vize_v)}", ulke: "{ulke_v}",'
            f' durum: "{map_durum(g(dc))}" as StatusType, durum_raw: "{g(dc)}",'
            f' statu: "{g(sc)}", evrakPct: "{g(evc)}", kaynak: "{g(kc)}",'
            f' not: "{g(ntc)}", danisman: "{sorumlu}", sehir: "{sehir}",'
            f' gorusme: "", takip: "", surec: "", karar: "", log: [], createdAt: "2026-02-25", updatedAt: "2026-02-25" }},'
        )
    return rows

ws1 = wb['Eray ESKİŞEHİR Lead']
eray_esk = read_sheet(ws1, 'Eray', 'Eskişehir', 1, 4, 5, 6, 8, 9, 10, 12, 13, 14)

ws2 = wb['Dilara ESKİŞEHİR Lead']
dilara = read_sheet(ws2, 'Dilara', 'Eskişehir', 1, 4, 5, 6, 8, 9, 10, 12, 13, 14)

ws3 = wb['Eray GAZİANTEPLead ']
eray_gaz = read_sheet(ws3, 'Eray', 'Gaziantep', 1, 4, 6, 7, 9, 10, 11, 13, 14, 15)

ws4 = wb['Elanur İSTANBULLead']
elanur = read_sheet(ws4, 'Elanur', 'İstanbul', 1, 3, 4, 5, 7, 8, 9, 11, 12, 13)

# Revenue data
for name in wb.sheetnames:
    if 'Gelir' in name:
        ws_gelir = wb[name]
        break

gelir_rows = []
for row in ws_gelir.iter_rows(min_row=4, max_row=ws_gelir.max_row, values_only=True):
    if not row[1]: continue
    def num(v):
        try:
            s = str(v).strip()
            if s in ['-', '', 'None']: return 0
            return float(s)
        except:
            return 0
    gid = str(int(num(row[0]))) if row[0] else str(len(gelir_rows)+1)
    gelir_rows.append(
        f'  {{ id: "{gid}", ad: "{clean(row[1])}", danisman: "{clean(row[2])}", sehir: "{clean(row[3])}",'
        f' odemeYontemi: "{clean(row[4])}", onOdemeTarihi: "{clean(row[5])}",'
        f' onOdeme: {num(row[6])}, kalanTarih: "{clean(row[7])}", kalanOdeme: {num(row[8])}, toplam: {num(row[9])} }},'
    )

out = [
    '// AUTO-GENERATED from Excel — DO NOT EDIT MANUALLY',
    'import { Customer, StatusType } from "../types";',
    '',
    'export const eskisehirData: Customer[] = [',
] + eray_esk + dilara + [
    '];',
    '',
    'export const gaziantepData: Customer[] = [',
] + eray_gaz + [
    '];',
    '',
    'export const istanbulData: Customer[] = [',
] + elanur + [
    '];',
    '',
    'export const allImportedCustomers: Customer[] = [...eskisehirData, ...gaziantepData, ...istanbulData];',
    '',
    'export interface RevenueEntry {',
    '  id: string;',
    '  ad: string;',
    '  danisman: string;',
    '  sehir: string;',
    '  odemeYontemi: string;',
    '  onOdemeTarihi: string;',
    '  onOdeme: number;',
    '  kalanTarih: string;',
    '  kalanOdeme: number;',
    '  toplam: number;',
    '}',
    '',
    'export const revenueData: RevenueEntry[] = [',
] + gelir_rows + [
    '];',
]

with open('C:/Users/Excalıbur/Desktop/vizemo_crm_app/src/data/importedData.ts', 'w', encoding='utf-8') as f:
    f.write('\n'.join(out))

print(f'OK: Esk={len(eray_esk)+len(dilara)}, Gaz={len(eray_gaz)}, Ist={len(elanur)}, Gelir={len(gelir_rows)}')
